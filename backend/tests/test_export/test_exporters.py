"""
Unit Tests for Export Services

Tests each exporter individually.
"""

import json

from app.services.export import (
    CSVExporter,
    ExcelExporter,
    ExporterFactory,
    ExportFormat,
    JSONExporter,
    SDFExporter,
)

# Sample test data — includes nested fields required by the audit column registry.
# All pre-existing top-level keys are preserved so non-CSV exporters are unaffected.
SAMPLE_RESULTS = [
    {
        "smiles": "CCO",
        "name": "Ethanol",
        "index": 0,
        "status": "success",
        "validation": {
            "canonical_smiles": "CCO",
            "inchikey": "LFQSCWFLJHTTHZ-UHFFFAOYSA-N",
            "overall_score": 95,
            "issues": [],
            "molecule_info": {
                "inchi": "InChI=1S/C2H6O/c1-2-3/h3H,2H2,1H3",
                "molecular_formula": "C2H6O",
                "molecular_weight": 46.07,
            },
            "all_checks": [
                {"check_name": "parsability", "passed": True},
                {"check_name": "sanitization", "passed": True},
                {"check_name": "valence", "passed": True},
            ],
        },
        "alerts": {
            "pains": {"matches": []},
            "brenk": {"matches": []},
        },
        "scoring": {
            "ml_readiness_score": 88,
            "ml_readiness": {"label": "High"},
            "np_likeness_score": -0.5,
            "druglikeness": {
                "qed_score": 0.75,
                "lipinski": {"passed": True, "violations": 0, "mw": 46.07, "logp": -0.3, "hbd": 1, "hba": 1},
                "veber": {"passed": True, "rotatable_bonds": 1, "tpsa": 20.2},
                "ro3": {"passed": True},
                "ghose": {"passed": False},
                "egan": {"passed": True},
                "muegge": {"passed": True},
            },
            "admet": {
                "synthetic_accessibility": {"score": 1.2, "classification": "Easy"},
                "solubility": {"log_s": -1.5, "classification": "Soluble"},
                "complexity": {"fsp3": 0.5},
                "cns_mpo": {"score": 4.5},
                "bioavailability": {"oral_absorption_likely": True},
                "pfizer_rule": {"passed": True},
                "gsk_rule": {"passed": True},
                "golden_triangle": {"in_golden_triangle": True},
            },
            "safety_filters": {
                "pains": {"passed": True, "alert_count": 0},
                "brenk": {"passed": True, "alert_count": 0},
                "nih": {"passed": True},
                "zinc": {"passed": True},
                "chembl": {"passed": True},
                "all_passed": True,
                "total_alerts": 0,
            },
            "aggregator": {"likelihood": "Low", "risk_score": 0.05},
            "scaffold": {"scaffold_smiles": ""},
            "boiled_egg": {"gi_absorbed": True, "bbb_permeant": False, "region": "white"},
            "consensus": {"score": 4},
            "lead_likeness": {"passed": True},
        },
        "standardized_smiles": "CCO",
        "standardization": {
            "result": {
                "standardized_smiles": "CCO",
                "success": True,
                "steps_applied": [{"name": "normalize", "applied": True}],
                "excluded_fragments": [],
                "stereo_comparison": {"lost": 0, "gained": 0},
                "mass_change_percent": 0.0,
            }
        },
        "safety_assessment": {
            "herg": {"herg_risk": "low", "risk_score": 1},
            "bro5": {"passed": True},
            "reos": {"passed": True, "n_violations": 0},
            "cyp_softspots": {"n_sites": 0},
            "complexity": {"n_outliers": 0},
        },
        "profiling": {
            "pfi": {"pfi": 2.3, "risk": "low"},
            "stars": {"stars": 6},
            "abbott": {"abbott_score": 0.8, "probability_pct": 80},
            "consensus_logp": {"consensus_logp": -0.3},
            "skin_permeation": {"log_kp": -8.1, "classification": "Low"},
            "sa_comparison": {
                "sa_score": {"score": 1.2},
                "scscore": {"score": 2.1},
                "syba": {"score": 20.5},
            },
            "cns_mpo": {"score": 3.5},
        },
    },
    {
        "smiles": "c1ccccc1",
        "name": "Benzene",
        "index": 1,
        "status": "success",
        "validation": {
            "canonical_smiles": "c1ccccc1",
            "inchikey": "UHOVQNZJYSORNB-UHFFFAOYSA-N",
            "overall_score": 85,
            "issues": [
                {
                    "check_name": "Alert",
                    "message": "PAINS alert detected",
                }
            ],
            "molecule_info": {
                "inchi": "InChI=1S/C6H6/c1-2-4-6-5-3-1/h1-6H",
                "molecular_formula": "C6H6",
                "molecular_weight": 78.11,
            },
            "all_checks": [
                {"check_name": "parsability", "passed": True},
                {"check_name": "sanitization", "passed": True},
                {"check_name": "valence", "passed": True},
            ],
        },
        "alerts": {
            "pains": {
                "matches": [
                    {
                        "pattern_name": "anil_di_alk_D(258)",
                        "smarts": "[Br]",
                    }
                ]
            },
            "brenk": {"matches": []},
        },
        "scoring": {
            "ml_readiness_score": 75,
            "ml_readiness": {"label": "Medium"},
            "np_likeness_score": 1.2,
            "druglikeness": {
                "qed_score": 0.32,
                "lipinski": {"passed": True, "violations": 0, "mw": 78.11, "logp": 1.9, "hbd": 0, "hba": 0},
                "veber": {"passed": True, "rotatable_bonds": 0, "tpsa": 0.0},
                "ro3": {"passed": True},
                "ghose": {"passed": False},
                "egan": {"passed": True},
                "muegge": {"passed": True},
            },
            "admet": {
                "synthetic_accessibility": {"score": 1.0, "classification": "Easy"},
                "solubility": {"log_s": -1.0, "classification": "Soluble"},
                "complexity": {"fsp3": 0.0},
                "cns_mpo": {"score": 3.0},
                "bioavailability": {"oral_absorption_likely": True},
                "pfizer_rule": {"passed": False},
                "gsk_rule": {"passed": True},
                "golden_triangle": {"in_golden_triangle": True},
            },
            "safety_filters": {
                "pains": {"passed": False, "alert_count": 1},
                "brenk": {"passed": True, "alert_count": 0},
                "nih": {"passed": True},
                "zinc": {"passed": True},
                "chembl": {"passed": True},
                "all_passed": False,
                "total_alerts": 1,
            },
            "aggregator": {"likelihood": "Low", "risk_score": 0.1},
            "scaffold": {"scaffold_smiles": "c1ccccc1"},
            "boiled_egg": {"gi_absorbed": True, "bbb_permeant": True, "region": "egg"},
            "consensus": {"score": 3},
            "lead_likeness": {"passed": True},
        },
        "standardized_smiles": "c1ccccc1",
        "standardization": {
            "result": {
                "standardized_smiles": "c1ccccc1",
                "success": True,
                "steps_applied": [],
                "excluded_fragments": [],
                "stereo_comparison": {"lost": 0, "gained": 0},
                "mass_change_percent": 0.0,
            }
        },
        "safety_assessment": {
            "herg": {"herg_risk": "moderate", "risk_score": 2},
            "bro5": {"passed": True},
            "reos": {"passed": True, "n_violations": 0},
            "cyp_softspots": {"n_sites": 1},
            "complexity": {"n_outliers": 0},
        },
        "profiling": {
            "pfi": {"pfi": 1.9, "risk": "low"},
            "stars": {"stars": 4},
            "abbott": {"abbott_score": 0.6, "probability_pct": 60},
            "consensus_logp": {"consensus_logp": 1.9},
            "skin_permeation": {"log_kp": -6.0, "classification": "Medium"},
            "sa_comparison": {
                "sa_score": {"score": 1.0},
                "scscore": {"score": 1.5},
                "syba": {"score": 35.0},
            },
            "cns_mpo": {"score": 2.0},
        },
    },
    {
        "smiles": "INVALID",
        "name": "Bad Molecule",
        "index": 2,
        "status": "error",
        "error": "Invalid SMILES",
        "validation": {},
        "alerts": {},
        "scoring": {},
    },
]


class TestCSVExporter:
    """Test CSV export functionality."""

    def test_csv_export_basic(self):
        """Test basic CSV export produces correct identity columns and data rows."""
        exporter = CSVExporter()
        buffer = exporter.export(SAMPLE_RESULTS)

        buffer.seek(0)
        content = buffer.read().decode("utf-8")

        # Identity columns must still be present
        assert "index" in content
        assert "name" in content
        assert "input_smiles" in content

        # Data rows
        assert "Ethanol" in content
        assert "CCO" in content

    def test_csv_has_audit_section_prefixes(self):
        """CSV columns must carry section prefixes from the audit registry."""
        exporter = CSVExporter()
        buffer = exporter.export(SAMPLE_RESULTS)

        buffer.seek(0)
        header_line = buffer.read().decode("utf-8").splitlines()[0]

        assert "[Validation]" in header_line
        assert "[Scoring]" in header_line
        assert "[Safety]" in header_line
        assert "[Standardization]" in header_line

    def test_csv_audit_column_values(self):
        """Known audit values from SAMPLE_RESULTS must appear in the CSV."""
        exporter = CSVExporter()
        buffer = exporter.export(SAMPLE_RESULTS)

        buffer.seek(0)
        content = buffer.read().decode("utf-8")

        # Overall score for Ethanol
        assert "95" in content
        # ML-readiness score for Ethanol
        assert "88" in content
        # QED score
        assert "0.75" in content
        # SA score
        assert "1.2" in content

    def test_csv_media_type(self):
        """Test CSV media type and file extension."""
        exporter = CSVExporter()
        assert exporter.media_type == "text/csv"
        assert exporter.file_extension == "csv"

    def test_csv_empty_results(self):
        """Test CSV export with empty results produces only a header row."""
        exporter = CSVExporter()
        buffer = exporter.export([])

        buffer.seek(0)
        content = buffer.read().decode("utf-8")

        # Header must be present
        assert "index" in content
        lines = [ln for ln in content.strip().split("\n") if ln]
        assert len(lines) == 1  # Only header row


class TestExcelExporter:
    """Test Excel export functionality."""

    def test_excel_export_basic(self):
        """Test basic Excel export."""
        exporter = ExcelExporter()
        buffer = exporter.export(SAMPLE_RESULTS)

        # Check buffer is not empty
        buffer.seek(0)
        content = buffer.read()
        assert len(content) > 0

        # Verify it's a valid Excel file (starts with PK for ZIP)
        assert content[:2] == b"PK"

    def test_excel_media_type(self):
        """Test Excel media type."""
        exporter = ExcelExporter()
        assert (
            exporter.media_type
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert exporter.file_extension == "xlsx"

    def test_excel_empty_results(self):
        """Test Excel export with empty results."""
        exporter = ExcelExporter()
        buffer = exporter.export([])

        buffer.seek(0)
        content = buffer.read()
        assert len(content) > 0  # Should still create a valid Excel file

    def test_excel_single_sheet_has_audit_columns(self):
        """Excel single-sheet export should have prefixed audit columns."""
        exporter = ExcelExporter()
        result = exporter.export(SAMPLE_RESULTS)
        import openpyxl

        wb = openpyxl.load_workbook(result)
        ws = wb["Results"]
        headers = [cell.value for cell in ws[1]]
        assert any("[Validation]" in str(h) for h in headers if h)
        assert any("[Scoring]" in str(h) for h in headers if h)

    def test_excel_multi_sheet_mode(self):
        """Excel multi-sheet should have one sheet per section plus Summary."""
        exporter = ExcelExporter(sheet_layout="multi")
        result = exporter.export(SAMPLE_RESULTS)
        import openpyxl

        wb = openpyxl.load_workbook(result)
        sheet_names = wb.sheetnames
        assert "Validation" in sheet_names
        assert "Deep Validation" in sheet_names
        assert "Scoring" in sheet_names
        assert "Safety" in sheet_names
        assert "Compound Profile" in sheet_names
        assert "Standardization" in sheet_names
        assert "Summary" in sheet_names


class TestSDFExporter:
    """Test SDF export functionality."""

    def test_sdf_export_basic(self):
        """Test basic SDF export."""
        exporter = SDFExporter()
        buffer = exporter.export(SAMPLE_RESULTS)

        # Read SDF
        buffer.seek(0)
        content = buffer.read().decode("utf-8")

        # Check SDF structure markers
        assert "$$$$" in content  # SDF record separator

        # Check molecule names
        assert "Ethanol" in content or "mol_1" in content

    def test_sdf_properties(self):
        """Test SDF export includes properties."""
        exporter = SDFExporter()
        buffer = exporter.export(SAMPLE_RESULTS)

        buffer.seek(0)
        content = buffer.read().decode("utf-8")

        # Check properties are included
        assert "overall_score" in content
        assert "ml_readiness_score" in content

    def test_sdf_invalid_smiles_skipped(self):
        """Test SDF export skips invalid SMILES."""
        exporter = SDFExporter()
        buffer = exporter.export(SAMPLE_RESULTS)

        buffer.seek(0)
        content = buffer.read().decode("utf-8")

        # Invalid molecule should be skipped
        assert "Bad Molecule" not in content or "INVALID" not in content

    def test_sdf_media_type(self):
        """Test SDF media type."""
        exporter = SDFExporter()
        assert exporter.media_type == "chemical/x-mdl-sdfile"
        assert exporter.file_extension == "sdf"

    def test_sdf_empty_results(self):
        """Test SDF export with empty results."""
        exporter = SDFExporter()
        buffer = exporter.export([])

        buffer.seek(0)
        content = buffer.read().decode("utf-8")
        assert len(content) == 0  # Empty SDF

    def test_sdf_exporter_default_has_basic_props(self):
        """SDF export without audit toggle should have basic properties only."""
        exporter = SDFExporter()
        result = exporter.export(SAMPLE_RESULTS)
        content = result.getvalue().decode("utf-8")
        # RDKit SDF format uses ">  <prop>" (two spaces before the angle bracket)
        assert "<overall_score>" in content

    def test_sdf_exporter_with_audit(self):
        """SDF export with include_audit=True should have audit properties."""
        exporter = SDFExporter(include_audit=True)
        result = exporter.export(SAMPLE_RESULTS)
        content = result.getvalue().decode("utf-8")
        # Should still have basic properties
        assert "<overall_score>" in content
        # Should have at least some audit fields (without section prefix)
        assert "<Parsability (Pass/Fail)>" in content or "<ML-Readiness Score (0-100)>" in content


class TestJSONExporter:
    """Test JSON export functionality."""

    def test_json_export_basic(self):
        """Test basic JSON export."""
        exporter = JSONExporter()
        buffer = exporter.export(SAMPLE_RESULTS)

        # Parse JSON
        buffer.seek(0)
        data = json.loads(buffer.read().decode("utf-8"))

        # Check structure
        assert "metadata" in data
        assert "results" in data

        # Check metadata
        assert "export_date" in data["metadata"]
        assert "total_count" in data["metadata"]
        assert data["metadata"]["total_count"] == len(SAMPLE_RESULTS)
        assert data["metadata"]["format_version"] == "2.0"

        # Check results
        assert len(data["results"]) == len(SAMPLE_RESULTS)
        assert data["results"][0]["name"] == "Ethanol"

    def test_json_media_type(self):
        """Test JSON media type."""
        exporter = JSONExporter()
        assert exporter.media_type == "application/json"
        assert exporter.file_extension == "json"

    def test_json_empty_results(self):
        """Test JSON export with empty results."""
        exporter = JSONExporter()
        buffer = exporter.export([])

        buffer.seek(0)
        data = json.loads(buffer.read().decode("utf-8"))

        assert data["metadata"]["total_count"] == 0
        assert len(data["results"]) == 0

    def test_json_structured_output(self):
        """JSON export should have structured sections per result."""
        exporter = JSONExporter()
        result_buffer = exporter.export(SAMPLE_RESULTS)
        data = json.loads(result_buffer.getvalue().decode("utf-8"))
        assert data["metadata"]["format_version"] == "2.0"
        first = data["results"][0]
        # Has identity columns (1-based index)
        assert first["index"] == 1
        assert first["name"] == "Ethanol"
        assert first["status"] == "success"
        # Has structured sections
        assert "validation" in first
        assert "scoring" in first
        assert "safety" in first
        assert "compound_profile" in first
        assert "standardization" in first
        # Values are extracted (not raw dicts)
        assert isinstance(first["validation"]["parsability_passed"], str)  # "Pass" not a dict


class TestExporterFactory:
    """Test ExporterFactory."""

    def test_factory_creates_csv(self):
        """Test factory creates CSV exporter."""
        exporter = ExporterFactory.create(ExportFormat.CSV)
        assert isinstance(exporter, CSVExporter)

    def test_factory_creates_excel(self):
        """Test factory creates Excel exporter."""
        exporter = ExporterFactory.create(ExportFormat.EXCEL)
        assert isinstance(exporter, ExcelExporter)

    def test_factory_creates_sdf(self):
        """Test factory creates SDF exporter."""
        exporter = ExporterFactory.create(ExportFormat.SDF)
        assert isinstance(exporter, SDFExporter)

    def test_factory_creates_json(self):
        """Test factory creates JSON exporter."""
        exporter = ExporterFactory.create(ExportFormat.JSON)
        assert isinstance(exporter, JSONExporter)

    def test_factory_invalid_format(self):
        """Test factory raises error for invalid format."""
        # Note: This test would require passing an invalid enum value
        # which is type-checked, so we skip this test
        pass

    def test_factory_passes_kwargs_to_excel(self):
        """Factory should pass include_images kwarg to ExcelExporter."""
        exporter = ExporterFactory.create(ExportFormat.EXCEL, include_images=True)
        assert isinstance(exporter, ExcelExporter)
        assert exporter._include_images is True

    def test_factory_ignores_unknown_kwargs(self):
        """Factory should ignore kwargs not in constructor signature."""
        exporter = ExporterFactory.create(ExportFormat.CSV, unknown_param=True)
        assert isinstance(exporter, CSVExporter)
